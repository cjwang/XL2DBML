import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('database.xlsx')
ws = wb['Tables']

# Create a dictionary of tables and their fields
tables = defaultdict(list)
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        current_table = row[0]
    tables[current_table].append(row[1:])

# Create a dictionary of referential constraints
referential_constraints = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[4]:
        referential_constraints[row[0] + '.' + row[1]] = row[4]

# Create a list of DBML statements
dbml_statements = []

# Create the DBML for each table
for table_name, fields in tables.items():
    # Create the DBML for this table
    dbml_table = f'Table {table_name} {{\n'
    dbml_fields = []
    dbml_primary_keys = []
    for field_name, data_type, primary_key in fields:
        dbml_fields.append(f'  {field_name} {data_type}{" [pk]" if primary_key else ""}')
        if primary_key:
            dbml_primary_keys.append(field_name)
    dbml_table += '\n'.join(dbml_fields) + '\n'
    if dbml_primary_keys:
        dbml_table += f'  pk ({", ".join(dbml_primary_keys)})\n'
    dbml_table += '}\n\n'
    dbml_statements.append(dbml_table)

# Create the DBML for the relationships
for field, reference in referential_constraints.items():
    if reference:
        if field.split('.')[0] != reference.split('.')[0]:
            dbml_statements.append(f'Ref: {field} < {reference}\n')
        else:
            dbml_statements.append(f'Ref: {field} = {reference}\n')

# Write the DBML to a file
with open('database.dbml', 'w') as f:
    f.writelines(dbml_statements)
